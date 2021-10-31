from datetime import date
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os



today = date.today()
a="16h"
x = f"il est {a + '-' + str(today)}"
print(x)
print()


path_import = {
    "P2K" : "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2K.xlsx",
    "P2J" : "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2J.xlsx",
    "P2N" : "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2N.xlsx"
}

path_export = {
    "P2K" : "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2K-1.xlsx",
    "P2J" : "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2J-1.xlsx",
    "P2N" : "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2N-1.xlsx"
}
#print(path_import["P2J"])

# def sap(env):
#     if len(env) > 1:
#         data = []
#         if len(env) == 2:
#             for i in env:
#                 print(i)
#                 df = pd.read_excel(path_import[i])
#                 data.append(df)

#                 pays = df['Affiliate'].unique()
                
#                 for sheet in pays:
#                     book = load_workbook(path_export[i])
#                     writer = pd.ExcelWriter(path_export[i], engine = 'openpyxl')
#                     writer.book = book

#                     d = df[df['Affiliate']==sheet]

#                     df.to_excel(writer, sheet_name = sheet, index=False)
#                     writer.save()
#                     writer.close()


#                 # if i == "P2K":
#                 #     df_P2K = pd.read_excel(path_import["P2K"])
#                 #     print(path_import["P2K"])
#                 #     print(df_P2K.head())
#                 # elif i == "P2J":
#                 #     df_P2J = pd.read_excel(path_import["P2J"])
#                 #     print(path_import["P2J"])
#                 #     print(df_P2J.head())
#                 # elif i == "P2N":
#                 #     df_P2N = pd.read_excel(path_import["P2N"])
#                 #     print(path_import["P2N"])
#                 #     print(df_P2N.head())

#             return data[0], data[1]

#         elif len(env) == 3:
#             for i in env:
#                 print(i)
#                 df = pd.read_excel(path_import[i])
#                 data.append(df)


#                 # if i == "P2K":
#                 #     df_P2K = pd.read_excel(path_import["P2K"])
#                 #     print(path_import["P2K"])
#                 #     print(df_P2K.head())
#                 # elif i == "P2J":
#                 #     df_P2J = pd.read_excel(path_import["P2J"])
#                 #     print(path_import["P2J"])
#                 #     print(df_P2J.head())
#                 # elif i == "P2N":
#                 #     df_P2N = pd.read_excel(path_import["P2N"])
#                 #     print(path_import["P2N"])
#                 #     print(df_P2N.head())

#             return data[0], data[1], data[3]
        
#     else:
#         data = pd.read_excel(path_import[env[0]])
#         print(env)

#         return data

# l = ["P2K", "P2K"]

# print(str(l[0]))

# def foo(s1):
#     return "'{}'".format(s1[0])

# s = foo(l)
# print(s)

# x = sap(l)


# print(x.shape)
# print(y.shape)

# def export(df):




# path_SAP_P2K = "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/SAP-P2J-1.xlsx"
# for sheet in pays_SAP_P2K:
#     book = load_workbook(path_SAP_P2K)
#     writer = pd.ExcelWriter(path_SAP_P2K, engine = 'openpyxl')
#     writer.book = book

#     df = SAP_P2K[SAP_P2K['Affiliate']==sheet]

#     df.to_excel(writer, sheet_name = sheet, index=False)
#     writer.save()
#     writer.close()

# d= pd.read_excel("C:/Users/J1049122/Desktop/Station Data/Master-Data/export/Ghana_2021-10-23.xlsx", sheet_name="SAP_vs_Sharepoint_SAPCode_BM")
# print(d.shape)
# print(d)

# for j in range(d.shape[0]):
#      print(d['SAPCode'][j])

# import os
# path = "C:/Users/J1049122/Desktop/Station Data/Master-Data/intermediaire/z"
# print(path)
# os.mkdir(path)

print("sharepoint")
sharepoint = pd.read_excel('C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/all-data-sharepoint.xlsx')
# sharepoint = sharepoint.iloc[:,:47]
print(sharepoint.columns)
# print(sharepoint.head())

print()

print("SAP")
sap = pd.read_excel('C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/Data-SAP.xlsx')
print(sap.columns)


dd = [   'Zone', 'SubZone', 'IntermediateStatus', 'Brand', 'Segment', 'ContractMode', 'ShopSegment', 'SFSActivity', 'SFSContractType', 'PartnerOrBrand', 'TargetKit', 'TargetPOSprovider', 
         'EstimatedInstallationDate', 'InstalledSolutionOnSite', 'SolutionProvider', 'SolutionInstallationDate', 'Status', 'SolutionRelease', 'SystemOwner', 'ConfigurationStatus',
         'IsAllPumpsConnectedToFCC', 'Reason', 'AutomaticTankGauging', 'ATGProvider', 'ATGModel', 'ATGConnected', 'ATGInstallationDate', 'TotalCardEPT connection', 'FuelCardProvider', 
         'EPTHardware', 'EPTModel', 'EPTNumber', 'EPTConnected', 'PaymentLocation', 'HOSInstalled', 'HOSProvider', 'WSMSoftwareInstalled', 'WSMProvider', 'TELECOM', 'STABILITE TELECOM',
         'STARTBOXStatus'
    ]

print()   
print(len(sharepoint.columns))
print(len(dd))




['SAPCode', 'Zone', 'SubZone', 'Affiliate', 'SAPName', 'Town',
       'IsActiveSite', 'IntermediateStatus', 'Brand', 'Segment',
       'BUSINESSMODEL', 'ContractMode', 'ShopSegment', 'SFSActivity',
       'SFSContractType', 'PartnerOrBrand', 'TargetKit', 'TargetPOSprovider',
       'EstimatedInstallationDate', 'InstalledSolutionOnSite',
       'SolutionProvider', 'SolutionInstallationDate', 'Status',
       'SolutionRelease', 'SystemOwner', 'ConfigurationStatus',
       'IsAllPumpsConnectedToFCC', 'Reason', 'AutomaticTankGauging',
       'ATGProvider', 'ATGModel', 'ATGConnected', 'ATGInstallationDate',
       'TotalCardEPT connection', 'FuelCardProvider', 'EPTHardware',
       'EPTModel', 'EPTNumber', 'EPTConnected', 'PaymentLocation',
       'HOSInstalled', 'HOSProvider', 'WSMSoftwareInstalled', 'WSMProvider',
       'TELECOM', 'STABILITE TELECOM', 'STARTBOXStatus']