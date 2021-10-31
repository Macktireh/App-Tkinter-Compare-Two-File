import pandas as pd
import numpy as np
import pyttsx3
import os
import shutil

from datetime import date
from openpyxl import load_workbook


# All Data (sharepoint)
sharepoint = pd.read_excel('C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/all-data-sharepoint.xlsx')
print('All Data SuiviSISDATA', sharepoint.shape)
# print(sharepoint.head())

# Data sharepoint APMO
SuiviSISDATA_APMO = sharepoint[sharepoint['Zone']=='APMO']
print('Data SuiviSISDATA APMO', SuiviSISDATA_APMO.shape)
# print(SuiviSISDATA_APMO.head())
pays_APMO = SuiviSISDATA_APMO['Affiliate'].unique()
print()
# print(pays_APMO)

# Data sharepoint AFR
SuiviSISDATA_AFR = sharepoint[sharepoint['Zone']=='AFR']
print('Data SuiviSISDATA AFR', SuiviSISDATA_AFR.shape)
# SuiviSISDATA_AFR.head()
pays_AFR = SuiviSISDATA_AFR['Affiliate'].unique()
print()
print(pays_AFR)

path_APMO = r"C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/sharepoint-APMO.xlsx"
path_AFR = r"C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/sharepoint-AFR.xlsx"

# if os.path.exists(path_APMO):
#     os.remove(path_APMO)
#     print(f"le fichier {path_APMO} à été bien supprimer\n-------------")
# else:
#     print(f"Impossible de supprimer le fichier {path_APMO} car il n'existe pas\n-------------")

# if os.path.exists(path_AFR):
#     os.remove(path_AFR)
#     print(f"le fichier {path_AFR} à été bien supprimer\n-------------")
# else:
#     print(f"Impossible de supprimer le fichier {path_AFR} car il n'existe pas\n-------------")


for sheet in pays_APMO:
    book = load_workbook(path_APMO)
    writer = pd.ExcelWriter(path_APMO, engine = 'openpyxl')
    writer.book = book

    df = SuiviSISDATA_APMO[SuiviSISDATA_APMO['Affiliate']==sheet]

    df.to_excel(writer, sheet_name = sheet, index=False)
    writer.save()
    writer.close()


for sheet in pays_AFR:
    book = load_workbook(path_AFR)
    writer = pd.ExcelWriter(path_AFR, engine = 'openpyxl')
    writer.book = book

    df = SuiviSISDATA_AFR[SuiviSISDATA_AFR['Affiliate']==sheet]

    df.to_excel(writer, sheet_name = sheet, index=False)
    writer.save()
    writer.close()

print()
print("--------------------")
print("Terminer avec succès")
print("--------------------")
print()